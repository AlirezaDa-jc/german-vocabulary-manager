"""
``StatisticsManager`` owns the Statistics sheet. It writes live Excel
formulas (never hard-coded numbers) so the counts stay correct as the
user adds, edits, or removes vocabulary — no script re-run required to
see updated numbers, only opening the file in Excel/LibreOffice.
"""

from __future__ import annotations

import datetime as _dt
import logging

from openpyxl import Workbook

import config

logger = logging.getLogger(__name__)


class StatisticsManager:
    """Writes and refreshes formulas on the Statistics sheet."""

    def write_formulas(self, workbook: Workbook) -> None:
        """Write the Statistics sheet formulas.

        Called once by ``create_excel.py`` right after all sheets and
        their headers exist. Uses COUNTIF/COUNTA against the Vocabulary
        sheet so every number recalculates live in Excel.
        """
        sheet = workbook[config.SHEET_STATS]
        vocab = config.SHEET_VOCAB
        vocab_cols = config.VOCAB_COLUMNS

        def _letter(idx: int) -> str:
            from openpyxl.utils import get_column_letter

            return get_column_letter(idx)

        german_letter = _letter(vocab_cols.index("German") + 1)
        wordtype_letter = _letter(vocab_cols.index("Word Type") + 1)
        favorite_letter = _letter(vocab_cols.index("Favorite") + 1)
        learned_letter = _letter(vocab_cols.index("Learned") + 1)

        rng = lambda letter: f"'{vocab}'!${letter}$2:${letter}$1048576"  # noqa: E731

        formulas = {
            "Total Words": f"=COUNTA({rng(german_letter)})",
            "Nouns": f'=COUNTIF({rng(wordtype_letter)},"Noun")',
            "Verbs": f'=COUNTIF({rng(wordtype_letter)},"Verb")',
            "Adjectives": f'=COUNTIF({rng(wordtype_letter)},"Adjective")',
            "Favorites": f'=COUNTIF({rng(favorite_letter)},"Yes")',
            "Learned": f'=COUNTIF({rng(learned_letter)},"Yes")',
        }

        label_col = 1
        value_col = 2
        row_by_label = {}
        for row_idx in range(2, sheet.max_row + 1):
            label = sheet.cell(row=row_idx, column=label_col).value
            if label:
                row_by_label[str(label).strip()] = row_idx

        for label, formula in formulas.items():
            row_idx = row_by_label.get(label)
            if row_idx is None:
                continue
            sheet.cell(row=row_idx, column=value_col, value=formula)

        progress_row = row_by_label.get("Progress %")
        total_row = row_by_label.get("Total Words")
        learned_row = row_by_label.get("Learned")
        if progress_row and total_row and learned_row:
            sheet.cell(
                row=progress_row,
                column=value_col,
                value=(
                    f"=IF(B{total_row}=0,0,B{learned_row}/B{total_row})"
                ),
            )
            sheet.cell(row=progress_row, column=value_col).number_format = "0.0%"

        logger.info("Statistics sheet formulas written/refreshed.")

    def touch_last_updated(self, workbook: Workbook) -> None:
        """Update the 'Last Updated' row on the Settings sheet."""
        sheet = workbook[config.SHEET_SETTINGS]
        now = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == "Last Updated":
                sheet.cell(row=row_idx, column=2, value=now)
                return
