"""
``ExcelManager`` is the single place that touches the workbook file at
runtime (``autofill.py``). It knows nothing about Wiktionary or any
other data source — it only knows how to read/write rows by column
name, which keeps ``autofill.py`` free of hard-coded column letters.

Conditional formatting (article colors, learned/favorite row fills) is
defined once in ``create_excel.py`` using formula-based rules, so
``ExcelManager`` never needs to paint cells manually — it only ever
writes values, and Excel recalculates the colors live.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Iterator, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import config

logger = logging.getLogger(__name__)


class ExcelManager:
    """Reads and writes the vocabulary workbook by column NAME, not letter."""

    def __init__(self, path: Path = config.WORKBOOK_PATH) -> None:
        self.path = path
        self.workbook: Optional[Workbook] = None
        self._header_index_cache: Dict[str, Dict[str, int]] = {}

    # ------------------------------------------------------------------
    # Lifecycle
    # ------------------------------------------------------------------

    def load(self) -> None:
        if not self.path.exists():
            raise FileNotFoundError(
                f"{self.path} does not exist. Run create_excel.py first."
            )
        self.workbook = load_workbook(self.path)
        self._header_index_cache.clear()

    def save(self) -> None:
        if self.workbook is None:
            raise RuntimeError("Workbook not loaded — call load() first.")
        self.workbook.save(self.path)
        logger.info("Workbook saved -> %s", self.path)

    # ------------------------------------------------------------------
    # Header helpers
    # ------------------------------------------------------------------

    def _sheet(self, sheet_name: str) -> Worksheet:
        if self.workbook is None:
            raise RuntimeError("Workbook not loaded — call load() first.")
        return self.workbook[sheet_name]

    def _header_map(self, sheet_name: str) -> Dict[str, int]:
        """Return ``{column_name: 1_based_column_index}`` for a sheet,
        reading the header row (row 1) once and caching the result.
        """
        if sheet_name in self._header_index_cache:
            return self._header_index_cache[sheet_name]
        sheet = self._sheet(sheet_name)
        headers: Dict[str, int] = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        self._header_index_cache[sheet_name] = headers
        return headers

    # ------------------------------------------------------------------
    # Word input sheet
    # ------------------------------------------------------------------

    def iter_pending_word_rows(self) -> Iterator[Tuple[int, str]]:
        """Yield ``(row_index, german_word)`` from the Word input sheet.

        A row is pending when it has a German word and its Processed cell
        is not "Yes".
        """
        sheet = self._sheet(config.SHEET_WORD)
        headers = self._header_map(config.SHEET_WORD)
        german_col = headers["German"]
        processed_col = headers["Processed"]

        for row_idx in range(2, sheet.max_row + 1):
            german_value = sheet.cell(row=row_idx, column=german_col).value
            if not german_value or not str(german_value).strip():
                continue

            processed_value = sheet.cell(row=row_idx, column=processed_col).value
            if processed_value and str(processed_value).strip().lower() == "yes":
                continue

            yield row_idx, str(german_value).strip()

    def update_word_row(self, row_idx: int, values: Dict[str, object]) -> None:
        """Write status values into the Word input sheet."""
        sheet = self._sheet(config.SHEET_WORD)
        headers = self._header_map(config.SHEET_WORD)

        for column_name, value in values.items():
            if value in (None, ""):
                continue
            col_idx = headers.get(column_name)
            if col_idx is None:
                logger.warning("Unknown Word column %r — skipped", column_name)
                continue
            sheet.cell(row=row_idx, column=col_idx, value=value)

    def append_vocab_seed_row(self, german_word: str) -> int:
        """Return an existing Vocabulary row for ``german_word`` or append one."""
        sheet = self._sheet(config.SHEET_VOCAB)
        headers = self._header_map(config.SHEET_VOCAB)
        german_col = headers["German"]

        first_empty_row: Optional[int] = None

        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=german_col).value

            if cell_value and str(cell_value).strip().lower() == german_word.strip().lower():
                return row_idx

            if first_empty_row is None and (cell_value is None or str(cell_value).strip() == ""):
                first_empty_row = row_idx

        target_row = first_empty_row or sheet.max_row + 1
        sheet.cell(row=target_row, column=german_col, value=german_word)
        return target_row

    # ------------------------------------------------------------------
    # Vocabulary sheet
    # ------------------------------------------------------------------

    def iter_pending_vocab_rows(self) -> Iterator[Tuple[int, str]]:
        """Yield ``(row_index, german_word)`` for every Vocabulary row that
        has a German word but has NOT been autofilled yet (its English
        cell is still empty). Already-completed rows are skipped so
        re-running the script is safe and fast.
        """
        sheet = self._sheet(config.SHEET_VOCAB)
        headers = self._header_map(config.SHEET_VOCAB)
        german_col = headers["German"]
        english_col = headers["English"]
        for row_idx in range(2, sheet.max_row + 1):
            german_value = sheet.cell(row=row_idx, column=german_col).value
            if not german_value or not str(german_value).strip():
                continue
            english_value = sheet.cell(row=row_idx, column=english_col).value
            if english_value and str(english_value).strip():
                continue  # already filled — skip
            yield row_idx, str(german_value).strip()

    def update_vocab_row(self, row_idx: int, values: Dict[str, object]) -> None:
        """Write ``{column_name: value}`` into ``row_idx`` of Vocabulary.

        Never overwrites a cell that already has a non-empty value
        (protects manual edits and the "Favorite"/"Learned"/"Review
        Date" columns the user maintains by hand).
        """
        sheet = self._sheet(config.SHEET_VOCAB)
        headers = self._header_map(config.SHEET_VOCAB)
        for column_name, value in values.items():
            if value in (None, ""):
                continue
            col_idx = headers.get(column_name)
            if col_idx is None:
                logger.warning("Unknown Vocabulary column %r — skipped", column_name)
                continue
            existing = sheet.cell(row=row_idx, column=col_idx).value
            if existing not in (None, ""):
                continue
            sheet.cell(row=row_idx, column=col_idx, value=value)

    # ------------------------------------------------------------------
    # Generic upsert for Verbs / Adjectives sheets (keyed by first column)
    # ------------------------------------------------------------------

    def upsert_row(
            self, sheet_name: str, key_column: str, key_value: str, values: Dict[str, object]
    ) -> int:
        """Insert or update a row in ``sheet_name`` keyed by ``key_column``.

        Returns the 1-based row index that was written.
        """
        sheet = self._sheet(sheet_name)
        headers = self._header_map(sheet_name)
        key_col_idx = headers[key_column]

        target_row: Optional[int] = None
        first_empty_row: Optional[int] = None

        for row_idx in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=key_col_idx).value

            if cell_value and str(cell_value).strip().lower() == key_value.strip().lower():
                target_row = row_idx
                break

            if first_empty_row is None and (cell_value is None or str(cell_value).strip() == ""):
                first_empty_row = row_idx

        if target_row is None:
            target_row = first_empty_row or sheet.max_row + 1
            sheet.cell(row=target_row, column=key_col_idx, value=key_value)

        for column_name, value in values.items():
            if value in (None, ""):
                continue
            col_idx = headers.get(column_name)
            if col_idx is None:
                logger.warning("Unknown %s column %r — skipped", sheet_name, column_name)
                continue
            existing = sheet.cell(row=target_row, column=col_idx).value
            if existing not in (None, ""):
                continue
            sheet.cell(row=target_row, column=col_idx, value=value)
        return target_row

    # ------------------------------------------------------------------
    # Convenience accessors used by StatisticsManager
    # ------------------------------------------------------------------

    def vocab_row_count(self) -> int:
        sheet = self._sheet(config.SHEET_VOCAB)
        headers = self._header_map(config.SHEET_VOCAB)
        german_col = headers["German"]
        count = 0
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=german_col).value:
                count += 1
        return count
