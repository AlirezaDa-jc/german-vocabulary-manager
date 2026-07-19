#!/usr/bin/env python3
"""
create_excel.py
=================

Generates ``vocabulary.xlsx`` from scratch using ``openpyxl``.

Run this once (or any time you want to reset the workbook):

    python create_excel.py

Creates six sheets: Vocabulary, Verbs, Adjectives, Grammar, Statistics,
Settings — with styled headers, filters, frozen header rows, table
formatting, data validation dropdowns, and conditional formatting
(article colors, learned/favorite row highlighting).
"""

from __future__ import annotations

import logging
from typing import List

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

import config
from core.statistics_manager import StatisticsManager

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

HEADER_FONT = Font(name=config.FONT_NAME, bold=True, color=config.HEADER_FONT_COLOR, size=11)
HEADER_FILL = PatternFill(
    start_color=config.HEADER_FILL_COLOR, end_color=config.HEADER_FILL_COLOR, fill_type="solid"
)
BODY_FONT = Font(name=config.FONT_NAME, size=11)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Generic sheet helpers
# ---------------------------------------------------------------------------


def _write_header(sheet: Worksheet, columns: List[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 24


def _autosize_columns(sheet: Worksheet, columns: List[str], min_width: int = 12, max_width: int = 40) -> None:
    for col_idx, name in enumerate(columns, start=1):
        width = max(min_width, min(max_width, len(name) + 6))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def _add_table(sheet: Worksheet, name: str, columns: List[str], n_data_rows: int = 200) -> None:
    """Register an Excel Table so filters + banded styling work out of the box."""
    last_col_letter = get_column_letter(len(columns))
    last_row = max(2, n_data_rows) + 1
    ref = f"A1:{last_col_letter}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showFirstColumn=False,
        showLastColumn=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _apply_body_style(sheet: Worksheet, columns: List[str], n_data_rows: int) -> None:
    for row_idx in range(2, n_data_rows + 2):
        for col_idx in range(1, len(columns) + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.alignment = LEFT_WRAP


# ---------------------------------------------------------------------------
# Word input sheet
# ---------------------------------------------------------------------------


def build_word_sheet(workbook: Workbook, n_data_rows: int = 200) -> None:
    sheet = workbook.create_sheet(config.SHEET_WORD)
    columns = config.WORD_COLUMNS
    _write_header(sheet, columns)
    _autosize_columns(sheet, columns)
    _apply_body_style(sheet, columns, n_data_rows)
    _add_table(sheet, "WordTable", columns, n_data_rows)

    last_row = n_data_rows + 1
    processed_col = columns.index("Processed") + 1
    processed_letter = get_column_letter(processed_col)

    dv_processed = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True,
        showDropDown=False,
    )
    sheet.add_data_validation(dv_processed)
    dv_processed.add(f"{processed_letter}2:{processed_letter}{last_row}")


# ---------------------------------------------------------------------------
# Vocabulary sheet
# ---------------------------------------------------------------------------


def build_vocabulary_sheet(workbook: Workbook, n_data_rows: int = 200) -> None:
    sheet = workbook.create_sheet(config.SHEET_VOCAB)
    columns = config.VOCAB_COLUMNS
    _write_header(sheet, columns)
    _autosize_columns(sheet, columns)
    _apply_body_style(sheet, columns, n_data_rows)
    _add_table(sheet, "VocabularyTable", columns, n_data_rows)

    last_row = n_data_rows + 1
    article_col = columns.index("Article") + 1
    wordtype_col = columns.index("Word Type") + 1
    favorite_col = columns.index("Favorite") + 1
    learned_col = columns.index("Learned") + 1
    review_col = columns.index("Review Date") + 1

    # --- Data validation dropdowns -----------------------------------
    article_letter = get_column_letter(article_col)
    wordtype_letter = get_column_letter(wordtype_col)
    favorite_letter = get_column_letter(favorite_col)
    learned_letter = get_column_letter(learned_col)

    dv_article = DataValidation(
        type="list", formula1='"der,die,das"', allow_blank=True, showDropDown=False
    )
    dv_wordtype = DataValidation(
        type="list",
        formula1=f'"{",".join(config.VALID_WORD_TYPES)}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_favorite = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
    dv_learned = DataValidation(
        type="list",
        formula1=f'"{config.LEARNED_UNCHECKED},{config.LEARNED_CHECKED}"',
        allow_blank=True,
        showDropDown=False,
    )

    for dv in (dv_article, dv_wordtype, dv_favorite, dv_learned):
        sheet.add_data_validation(dv)

    dv_article.add(f"{article_letter}2:{article_letter}{last_row}")
    dv_wordtype.add(f"{wordtype_letter}2:{wordtype_letter}{last_row}")
    dv_favorite.add(f"{favorite_letter}2:{favorite_letter}{last_row}")
    dv_learned.add(f"{learned_letter}2:{learned_letter}{last_row}")

    for row_idx in range(2, last_row + 1):
        sheet.cell(row=row_idx, column=learned_col, value=config.LEARNED_UNCHECKED)
        review_cell = sheet.cell(row=row_idx, column=review_col)
        review_cell.value = f'=IF(${learned_letter}{row_idx}="{config.LEARNED_CHECKED}",TODAY(),"")'
        review_cell.number_format = "yyyy-mm-dd"

    # --- Conditional formatting ---------------------------------------
    full_row_range = f"A2:{get_column_letter(len(columns))}{last_row}"

    # 1) Article text color (applies to the Article cell itself)
    article_range = f"{article_letter}2:{article_letter}{last_row}"
    sheet.conditional_formatting.add(
        article_range,
        FormulaRule(
            formula=[f'${article_letter}2="der"'],
            font=Font(color=config.DER_COLOR, bold=True),
        ),
    )
    sheet.conditional_formatting.add(
        article_range,
        FormulaRule(
            formula=[f'${article_letter}2="die"'],
            font=Font(color=config.DIE_COLOR, bold=True),
        ),
    )
    sheet.conditional_formatting.add(
        article_range,
        FormulaRule(
            formula=[f'${article_letter}2="das"'],
            font=Font(color=config.DAS_COLOR, bold=True),
        ),
    )

    # 2) Learned rows -> green fill when selected as checked
    learned_fill = PatternFill(
        start_color=config.LEARNED_FILL, end_color=config.LEARNED_FILL, fill_type="solid"
    )
    learned_rule = FormulaRule(
        formula=[f'${learned_letter}2="{config.LEARNED_CHECKED}"'],
        fill=learned_fill,
        stopIfTrue=True,
    )
    sheet.conditional_formatting.add(full_row_range, learned_rule)

    # 3) Favorite rows -> yellow fill (only applies if not already Learned)
    favorite_fill = PatternFill(
        start_color=config.FAVORITE_FILL, end_color=config.FAVORITE_FILL, fill_type="solid"
    )
    favorite_rule = FormulaRule(
        formula=[f'${favorite_letter}2="Yes"'], fill=favorite_fill, stopIfTrue=True
    )
    sheet.conditional_formatting.add(full_row_range, favorite_rule)

    # Review Date column: widen + date format hint
    sheet.column_dimensions[get_column_letter(review_col)].width = 16
    sheet.column_dimensions[learned_letter].width = 18


# ---------------------------------------------------------------------------
# Verbs sheet
# ---------------------------------------------------------------------------


def build_verbs_sheet(workbook: Workbook, n_data_rows: int = 200) -> None:
    sheet = workbook.create_sheet(config.SHEET_VERBS)
    columns = config.VERB_COLUMNS
    _write_header(sheet, columns)
    _autosize_columns(sheet, columns)
    _apply_body_style(sheet, columns, n_data_rows)
    _add_table(sheet, "VerbsTable", columns, n_data_rows)

    last_row = n_data_rows + 1
    for name in ("Hilfsverb", "Trennbar", "Reflexiv", "Irregular"):
        col_idx = columns.index(name) + 1
        letter = get_column_letter(col_idx)
        if name == "Hilfsverb":
            dv = DataValidation(type="list", formula1='"haben,sein"', allow_blank=True)
        else:
            dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
        sheet.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last_row}")


# ---------------------------------------------------------------------------
# Adjectives sheet
# ---------------------------------------------------------------------------


def build_adjectives_sheet(workbook: Workbook, n_data_rows: int = 200) -> None:
    sheet = workbook.create_sheet(config.SHEET_ADJECTIVES)
    columns = config.ADJECTIVE_COLUMNS
    _write_header(sheet, columns)
    _autosize_columns(sheet, columns)
    _apply_body_style(sheet, columns, n_data_rows)
    _add_table(sheet, "AdjectivesTable", columns, n_data_rows)


# ---------------------------------------------------------------------------
# Grammar sheet (static reference content, written once)
# ---------------------------------------------------------------------------

_GRAMMAR_CONTENT = [
    (
        "Noun Gender & Articles",
        "Every German noun has a fixed grammatical gender: der (masculine), "
        "die (feminine), das (neuter). The article changes across the four "
        "cases (Nominativ, Genitiv, Dativ, Akkusativ).",
        "der Mann / die Frau / das Kind",
    ),
    (
        "Plural Formation",
        "German plurals are irregular and must be memorized per noun; common "
        "patterns add -e, -er, -(e)n, -s, or use an umlaut with no ending.",
        "das Haus -> die Häuser; die Blume -> die Blumen",
    ),
    (
        "The Four Cases",
        "Nominativ = subject, Akkusativ = direct object, Dativ = indirect "
        "object, Genitiv = possession. Articles and adjective endings change "
        "with the case.",
        "Der Hund (Nom.) beißt den Mann (Akk.).",
    ),
    (
        "Present Tense Conjugation",
        "Regular (weak) verbs conjugate by adding -e/-st/-t/-en/-t/-en to the "
        "stem for ich/du/er-sie-es/wir/ihr/sie-Sie.",
        "machen: ich mache, du machst, er macht, wir machen, ihr macht, sie machen",
    ),
    (
        "Separable Verbs (trennbare Verben)",
        "Verbs with a separable prefix (ab-, an-, auf-, aus-, mit-, ...) split "
        "the prefix to the end of the clause in main-clause present/past tense.",
        "aufstehen -> Ich stehe um sieben Uhr auf.",
    ),
    (
        "Perfekt Tense",
        "Formed with haben/sein (conjugated) + Partizip II at the end of the "
        "clause. Most verbs use haben; motion/change-of-state verbs use sein.",
        "Ich habe gegessen. / Ich bin gegangen.",
    ),
    (
        "Adjective Comparison",
        "Comparative adds -er, superlative uses am ...(e)sten; several common "
        "adjectives are irregular.",
        "schön -> schöner -> am schönsten; gut -> besser -> am besten",
    ),
    (
        "Word Order in Subordinate Clauses",
        "In subordinate clauses (introduced by weil, dass, wenn, ...) the "
        "conjugated verb moves to the very end of the clause.",
        "Ich bleibe zu Hause, weil es regnet.",
    ),
]


def build_grammar_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(config.SHEET_GRAMMAR)
    columns = config.GRAMMAR_COLUMNS
    _write_header(sheet, columns)
    for row_idx, (topic, rule, example) in enumerate(_GRAMMAR_CONTENT, start=2):
        sheet.cell(row=row_idx, column=1, value=topic).font = Font(
            name=config.FONT_NAME, bold=True
        )
        sheet.cell(row=row_idx, column=2, value=rule).alignment = LEFT_WRAP
        sheet.cell(row=row_idx, column=3, value=example).alignment = LEFT_WRAP
        sheet.cell(row=row_idx, column=2).font = BODY_FONT
        sheet.cell(row=row_idx, column=3).font = Font(
            name=config.FONT_NAME, italic=True
        )
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 70
    sheet.column_dimensions["C"].width = 45
    for row_idx in range(2, len(_GRAMMAR_CONTENT) + 2):
        sheet.row_dimensions[row_idx].height = 45
    sheet.auto_filter.ref = f"A1:C{len(_GRAMMAR_CONTENT) + 1}"


# ---------------------------------------------------------------------------
# Statistics sheet
# ---------------------------------------------------------------------------


def build_statistics_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(config.SHEET_STATS)
    _write_header(sheet, ["Metric", "Value"])
    for row_idx, label in enumerate(config.STATS_ROWS, start=2):
        cell = sheet.cell(row=row_idx, column=1, value=label)
        cell.font = Font(name=config.FONT_NAME, bold=True)
        sheet.cell(row=row_idx, column=2).font = BODY_FONT
        sheet.cell(row=row_idx, column=2).alignment = CENTER
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 18
    StatisticsManager().write_formulas(workbook)


# ---------------------------------------------------------------------------
# Settings sheet
# ---------------------------------------------------------------------------


def build_settings_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(config.SHEET_SETTINGS)
    _write_header(sheet, ["Setting", "Value"])
    for row_idx, (key, value) in enumerate(config.SETTINGS_ROWS, start=2):
        sheet.cell(row=row_idx, column=1, value=key).font = Font(
            name=config.FONT_NAME, bold=True
        )
        sheet.cell(row=row_idx, column=2, value=value).font = BODY_FONT
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 55


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)  # drop the default "Sheet"

    build_word_sheet(workbook)
    build_vocabulary_sheet(workbook)
    build_verbs_sheet(workbook)
    build_adjectives_sheet(workbook)
    build_grammar_sheet(workbook)
    build_statistics_sheet(workbook)
    build_settings_sheet(workbook)

    workbook.save(config.WORKBOOK_PATH)
    logger.info("Created %s", config.WORKBOOK_PATH)


if __name__ == "__main__":
    main()
