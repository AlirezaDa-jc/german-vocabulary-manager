"""
core/workbook_viewer.py
=========================

A simple, formula-free, in-app spreadsheet viewer/editor. Opens the
workbook in a new window with one tab per sheet, using ``tksheet`` grids.
No Excel installation is required.

- Conditional formatting (article colors, Learned/Favorite row colors)
  is re-applied in Python at render time — a lightweight approximation
  of the real Excel conditional-formatting rules, not the rules
  themselves.
- Saving writes edited cell values back into the SAME in-memory
  ``openpyxl`` Workbook object that was loaded from disk (not a fresh
  rebuild), so anything this viewer doesn't touch — Excel Tables, data
  validation, conditional formatting rules, column widths, etc. — is
  left completely intact when the file is written back out.
- Double-click a "Pronunciation URL" cell to play that word's audio
  with the OS's default handler for that file type.
- Select a row on the Vocabulary tab and click "Toggle Learned" to
  flip its Learned status; Review Date is set to today when checked,
  cleared when unchecked.
"""

from __future__ import annotations

import datetime as _dt
import logging
import os
import platform
import subprocess
import tkinter as tk
from pathlib import Path
from tkinter import messagebox, ttk
from typing import Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from tksheet import Sheet

import config

logger = logging.getLogger(__name__)

# Approximate colors matching create_excel.py's conditional formatting
_DER_COLOR = "#1E3A8A"
_DIE_COLOR = "#991B1B"
_DAS_COLOR = "#166534"
_LEARNED_FILL = "#C6EFCE"
_FAVORITE_FILL = "#FFF2CC"

# Assumed sheet name for the user-input sheet — change this if your
# config.py defines a different constant (e.g. config.SHEET_WORD).
_SHEET_WORD = "Word"


class WorkbookViewerWindow(tk.Toplevel):
    """A Toplevel window showing every sheet of the workbook as an
    editable grid, with Save, Toggle Learned, and double-click-to-play
    audio.
    """

    def __init__(
        self, master: tk.Misc, workbook_path: Path = config.WORKBOOK_PATH
    ) -> None:
        super().__init__(master)
        self.title(f"Vocabulary Workbook — {workbook_path.name}")
        self.geometry("1200x700")

        self.workbook_path = workbook_path
        self.workbook: Optional[Workbook] = None
        self._sheets: Dict[str, Sheet] = {}

        self._build_toolbar()
        self._notebook = ttk.Notebook(self)
        self._notebook.pack(fill="both", expand=True)
        self._notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self._status = tk.StringVar(value="Loading...")
        ttk.Label(self, textvariable=self._status, anchor="w").pack(
            fill="x", padx=6, pady=(0, 4)
        )

        self.load()

    # ------------------------------------------------------------------
    # Toolbar
    # ------------------------------------------------------------------

    def _build_toolbar(self) -> None:
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=6, pady=6)
        ttk.Button(bar, text="💾 Save", command=self.save).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(bar, text="🔄 Reload", command=self.load).pack(
            side="left", padx=(0, 6)
        )
        self.toggle_learned_button = ttk.Button(
            bar, text="☑️ Toggle Learned", command=self.toggle_learned, state="disabled"
        )
        self.toggle_learned_button.pack(side="left", padx=(0, 6))
        self.new_row_button = ttk.Button(
            bar, text="➕ New Row", command=self.add_new_word_row, state="disabled"
        )
        self.new_row_button.pack(side="left", padx=(0, 6))
        ttk.Label(
            bar,
            text="Toggle Learned works on Vocabulary. New Row works on Word. Double-click 'Pronunciation URL' to play audio.",
            foreground="#555555",
        ).pack(side="left", padx=12)

    def _on_tab_changed(self, event=None) -> None:
        try:
            current_tab = self._notebook.tab(self._notebook.select(), "text")
        except tk.TclError:
            return
        self.toggle_learned_button.configure(
            state="normal" if current_tab == config.SHEET_VOCAB else "disabled"
        )
        self.new_row_button.configure(
            state="normal" if current_tab == _SHEET_WORD else "disabled"
        )

    # ------------------------------------------------------------------
    # Loading
    # ------------------------------------------------------------------

    def load(self) -> None:
        if not self.workbook_path.exists():
            messagebox.showerror(
                "Workbook not found",
                f"{self.workbook_path} does not exist. Create it first.",
            )
            self.destroy()
            return

        for tab_id in self._notebook.tabs():
            self._notebook.forget(tab_id)
        self._sheets.clear()

        try:
            self.workbook = load_workbook(self.workbook_path)
        except PermissionError:
            messagebox.showerror(
                "Workbook is open elsewhere",
                "Close the workbook in Excel/Numbers before opening it here.",
            )
            self.destroy()
            return

        for sheet_name in self.workbook.sheetnames:
            self._add_sheet_tab(sheet_name)

        self._on_tab_changed()
        self._status.set(f"Loaded {self.workbook_path}")

    def _add_sheet_tab(self, sheet_name: str) -> None:
        worksheet: Worksheet = self.workbook[sheet_name]
        frame = ttk.Frame(self._notebook)
        self._notebook.add(frame, text=sheet_name)

        headers, rows = self._read_sheet_values(worksheet)

        sheet_widget = Sheet(
            frame,
            data=rows,
            headers=headers,
            show_row_index=True,
        )
        sheet_widget.enable_bindings(
            "single_select",
            "row_select",
            "column_select",
            "arrowkeys",
            "edit_cell",
            "copy",
            "paste",
            "delete",
            "undo",
        )
        sheet_widget.pack(fill="both", expand=True, padx=4, pady=4)

        pronunciation_col = None
        if "Pronunciation URL" in headers:
            pronunciation_col = headers.index("Pronunciation URL")
        sheet_widget.extra_bindings(
            "cell_select",
            func=lambda event, col=pronunciation_col, name=sheet_name: self._maybe_play(
                event, col, name
            ),
        )

        if sheet_name == config.SHEET_VOCAB:
            self._apply_vocab_highlighting(sheet_widget, headers, rows)

        self._sheets[sheet_name] = sheet_widget

    def _read_sheet_values(self, worksheet: Worksheet):
        headers: List[str] = [cell.value or "" for cell in worksheet[1]]
        rows: List[List[str]] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row and any(v not in (None, "") for v in row):
                rows.append(["" if v is None else v for v in row])

        if worksheet.title == config.SHEET_STATS:
            headers, rows = self._compute_statistics_snapshot()

        return headers, rows

    def _compute_statistics_snapshot(self):
        """Recompute the Statistics numbers in plain Python (no formulas)."""
        vocab = self.workbook[config.SHEET_VOCAB]
        vocab_headers = {cell.value: idx for idx, cell in enumerate(vocab[1])}
        values = list(vocab.iter_rows(min_row=2, values_only=True))

        def count_col(name: str, match: Optional[str] = None) -> int:
            idx = vocab_headers.get(name)
            if idx is None:
                return 0
            if match is None:
                return sum(
                    1 for row in values if idx < len(row) and row[idx] not in (None, "")
                )
            return sum(1 for row in values if idx < len(row) and row[idx] == match)

        total = count_col("German")
        nouns = count_col("Word Type", "Noun")
        verbs = count_col("Word Type", "Verb")
        adjectives = count_col("Word Type", "Adjective")
        favorites = count_col("Favorite", "Yes")
        learned = count_col("Learned", config.LEARNED_CHECKED)
        progress = f"{(learned / total * 100):.1f}%" if total else "0.0%"

        headers = ["Metric", "Value"]
        rows = [
            ["Total Words", total],
            ["Nouns", nouns],
            ["Verbs", verbs],
            ["Adjectives", adjectives],
            ["Favorites", favorites],
            ["Learned", learned],
            ["Progress %", progress],
        ]
        return headers, rows

    def _apply_vocab_highlighting(
        self, sheet_widget: Sheet, headers: List[str], rows: List[List]
    ) -> None:
        """Lightweight, render-time approximation of the Excel
        conditional formatting rules (article colors, Learned/Favorite
        row fills). Recomputed every time it's called — not a stored rule.
        """
        try:
            article_col = headers.index("Article")
            learned_col = headers.index("Learned")
            favorite_col = headers.index("Favorite")
        except ValueError:
            return

        sheet_widget.dehighlight_all()

        for row_idx, row in enumerate(rows):
            article = row[article_col] if article_col < len(row) else None
            color = {"der": _DER_COLOR, "die": _DIE_COLOR, "das": _DAS_COLOR}.get(
                article
            )
            if color:
                sheet_widget.highlight_cells(
                    row=row_idx, column=article_col, fg=color, redraw=False
                )

            learned = row[learned_col] if learned_col < len(row) else None
            favorite = row[favorite_col] if favorite_col < len(row) else None
            fill = None
            if learned == config.LEARNED_CHECKED:
                fill = _LEARNED_FILL
            elif favorite == "Yes":
                fill = _FAVORITE_FILL
            if fill:
                sheet_widget.highlight_rows(rows=[row_idx], bg=fill, redraw=False)

        sheet_widget.redraw()

    # ------------------------------------------------------------------
    # Toggle Learned
    # ------------------------------------------------------------------

    def toggle_learned(self) -> None:
        sheet_widget = self._sheets.get(config.SHEET_VOCAB)
        if sheet_widget is None:
            return

        selected = sheet_widget.get_currently_selected()
        if not selected:
            messagebox.showinfo(
                "Toggle Learned", "Click a row on the Vocabulary sheet first."
            )
            return
        row_idx = selected.row

        headers = sheet_widget.headers()
        if "Learned" not in headers:
            return
        learned_col = headers.index("Learned")
        current = sheet_widget.get_cell_data(row_idx, learned_col)
        now_checked = current != config.LEARNED_CHECKED
        new_learned_value = (
            config.LEARNED_CHECKED if now_checked else config.LEARNED_UNCHECKED
        )
        sheet_widget.set_cell_data(
            row_idx, learned_col, new_learned_value, redraw=False
        )

        if "Review Date" in headers:
            review_col = headers.index("Review Date")
            new_review_value = _dt.date.today().isoformat() if now_checked else ""
            sheet_widget.set_cell_data(
                row_idx, review_col, new_review_value, redraw=False
            )

        rows = sheet_widget.get_sheet_data()
        self._apply_vocab_highlighting(sheet_widget, headers, rows)

    # ------------------------------------------------------------------
    # New Row (Word sheet)
    # ------------------------------------------------------------------

    def add_new_word_row(self) -> None:
        sheet_widget = self._sheets.get(_SHEET_WORD)
        if sheet_widget is None:
            messagebox.showwarning("New Row", f"'{_SHEET_WORD}' sheet not found.")
            return

        sheet_widget.insert_row()
        new_row_idx = len(sheet_widget.get_sheet_data()) - 1
        sheet_widget.set_currently_selected(new_row_idx, 0)
        sheet_widget.see(row=new_row_idx, column=0)

    # ------------------------------------------------------------------
    # Play audio
    # ------------------------------------------------------------------

    def _maybe_play(
        self, event, pronunciation_col: Optional[int], sheet_name: str
    ) -> None:
        if pronunciation_col is None:
            return
        selected = getattr(event, "selected", None)
        if not selected or selected.row is None or selected.column is None:
            return
        row_idx, col_idx = selected.row, selected.column
        if col_idx != pronunciation_col:
            return

        sheet_widget = self._sheets[sheet_name]
        value = sheet_widget.get_cell_data(row_idx, col_idx)
        if not value:
            return

        audio_path = (self.workbook_path.parent / str(value)).resolve()
        if not audio_path.exists():
            messagebox.showwarning("Audio not found", f"File not found:\n{audio_path}")
            return
        self._play_audio(audio_path)

    @staticmethod
    def _play_audio(path: Path) -> None:
        """Launch the OS default handler for the audio file.

        Note: on stock Windows, .ogg files need a player that supports
        Ogg Vorbis (e.g. VLC) registered as the default handler — the
        Windows-bundled player alone may not decode .ogg.
        """
        system = platform.system()
        try:
            if system == "Windows":
                os.startfile(str(path))  # type: ignore[attr-defined]
            elif system == "Darwin":
                subprocess.run(["open", str(path)], check=False)
            else:
                subprocess.run(["xdg-open", str(path)], check=False)
        except OSError as exc:
            logger.error("Could not play audio %s: %s", path, exc)
            messagebox.showerror("Playback failed", str(exc))

    # ------------------------------------------------------------------
    # Saving
    # ------------------------------------------------------------------

    def save(self) -> None:
        if self.workbook is None:
            return

        for sheet_name, sheet_widget in self._sheets.items():
            if sheet_name == config.SHEET_STATS:
                continue  # computed, not user-editable data
            worksheet = self.workbook[sheet_name]
            headers = [cell.value for cell in worksheet[1]]
            data = sheet_widget.get_sheet_data()
            for row_offset, row_values in enumerate(data):
                excel_row = row_offset + 2  # header is row 1
                for col_idx, value in enumerate(row_values):
                    if col_idx >= len(headers):
                        continue
                    cell = worksheet.cell(row=excel_row, column=col_idx + 1)
                    cell.value = value if value != "" else None

        try:
            self.workbook.save(self.workbook_path)
        except PermissionError:
            messagebox.showerror(
                "Cannot save",
                "The workbook is open in another program (Excel/Numbers). "
                "Close it there first, then Save again.",
            )
            return

        self._status.set(f"Saved {self.workbook_path}")
        messagebox.showinfo("Saved", "Workbook saved successfully.")


def open_workbook_viewer(
    master: tk.Misc, workbook_path: Path = config.WORKBOOK_PATH
) -> None:
    """Convenience entry point to call from app.py's button command."""
    WorkbookViewerWindow(master, workbook_path)
